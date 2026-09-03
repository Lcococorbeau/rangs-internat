#!/usr/bin/env python3
"""Convert EDN simulation source files into the JSON format used by the site.

Supported layout:
    sources/<year>/Tour 1.xlsx
    sources/<year>/Tour 2.csv
    ...
    sources/<year>/Tour Def.xlsx

XLSX format:
- one worksheet per specialty;
- worksheet order is irrelevant;
- specialty names are read from worksheet names (known truncated Excel sheet names
  are mapped back to their canonical names);
- expected columns include a city/subdivision field and a maximum-rank field.

CSV format:
- one file represents one simulation round;
- it MUST contain a specialty column because CSV has no worksheets;
- it also needs city/subdivision and maximum-rank columns.

The importer is intentionally strict when a file is ambiguous: an invalid upload must
fail the build rather than silently publish incorrect data.
"""
from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import OrderedDict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCES = ROOT / "sources"
DATA_DIR = ROOT / "data"
SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}
HEADER_SCAN_ROWS = 12

# Canonical display names for specialties already encountered in EDN files.
# This is an alias registry only: the importer does NOT require this exact list,
# count, or order. Unknown/new specialties are accepted automatically.
KNOWN_SPECIALTIES = [
    "Allergologie", "Anatomie et cytologie pathologiques", "Anesthésie-réanimation",
    "Biologie médicale", "Chirurgie maxillo-faciale", "Chirurgie orale",
    "Chirurgie orthopédique et traumatologique", "Chirurgie plastique, reconstructrice et esthétique",
    "Chirurgie pédiatrique", "Chirurgie thoracique et cardiovasculaire", "Chirurgie vasculaire",
    "Chirurgie viscérale et digestive", "Dermatologie et vénéréologie",
    "Endocrinologie-diabétologie-nutrition", "Gynécologie médicale", "Gynécologie obstétrique",
    "Génétique médicale", "Gériatrie", "Hématologie", "Hépato-gastro-entérologie",
    "Maladies infectieuses et tropicales", "Médecine cardiovasculaire", "Médecine d’urgence",
    "Médecine et santé au travail", "Médecine générale", "Médecine intensive-réanimation",
    "Médecine interne et immunologie clinique", "Médecine légale et expertises médicales",
    "Médecine nucléaire", "Médecine physique et de réadaptation", "Médecine vasculaire",
    "Neurochirurgie", "Neurologie", "Néphrologie", "Oncologie", "Ophtalmologie",
    "Oto-rhino-laryngologie - chirurgie cervico-faciale", "Pneumologie", "Psychiatrie", "Pédiatrie",
    "Radiologie et imagerie médicale", "Rhumatologie", "Santé publique", "Urologie",
]


def norm(value: Any) -> str:
    """Accent/case/punctuation-insensitive normalization for matching labels."""
    s = "" if value is None else str(value)
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"[’'`´]", "'", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


KNOWN_BY_NORM = {norm(name): name for name in KNOWN_SPECIALTIES}


def canonical_specialty(raw: Any) -> str:
    """Return a stable specialty display name without relying on worksheet order."""
    name = re.sub(r"\s+", " ", str(raw or "").strip())
    if not name:
        raise ValueError("Nom de spécialité vide.")
    key = norm(name)
    if key in KNOWN_BY_NORM:
        return KNOWN_BY_NORM[key]

    # Excel worksheet names are limited to 31 characters. Match a truncated
    # worksheet title only when it identifies exactly one known specialty.
    candidates = []
    for canonical in KNOWN_SPECIALTIES:
        ck = norm(canonical)
        if len(key) >= 8 and (ck.startswith(key) or key.startswith(ck)):
            candidates.append(canonical)
    if len(candidates) == 1:
        return candidates[0]

    # New specialties are deliberately accepted. Their worksheet/CSV spelling
    # becomes the display name for that year.
    return name


def canonical_city(raw: Any) -> str:
    """Normalize subdivision labels so minor casing/accent differences do not split curves."""
    text = re.sub(r"\s+", " ", str(raw or "").strip())
    if not text:
        raise ValueError("Ville/subdivision vide.")
    key = norm(text)
    aliases = {
        "paris": "AP-HP", "ap hp": "AP-HP", "aphp": "AP-HP", "paris ap hp": "AP-HP",
        "marseille": "AP-HM", "ap hm": "AP-HM", "aphm": "AP-HM", "marseille ap hm": "AP-HM",
        "lyon": "HCL", "hcl": "HCL", "hospices civils de lyon": "HCL", "lyon hcl": "HCL",
        "martinique guadeloupe": "MARTINIQUE / GUADELOUPE",
        "guadeloupe martinique": "MARTINIQUE / GUADELOUPE",
        "la reunion": "LA REUNION", "reunion": "LA REUNION",
    }
    if key in aliases:
        return aliases[key]
    # Historical files use uppercase, mostly accent-free subdivision labels.
    ascii_text = unicodedata.normalize("NFD", text)
    ascii_text = "".join(ch for ch in ascii_text if unicodedata.category(ch) != "Mn")
    ascii_text = re.sub(r"\s*[-–—]\s*", "-", ascii_text)
    ascii_text = re.sub(r"\s*/\s*", " / ", ascii_text)
    return re.sub(r"\s+", " ", ascii_text).strip().upper()


def tour_from_filename(path: Path) -> tuple[str, int]:
    name = norm(path.stem)
    if re.search(r"\b(def|definitif|final|finale)\b", name):
        return "Définitif", 10_000
    patterns = [r"\btour\s*(\d+)\b", r"\bt\s*(\d+)\b"]
    for pattern in patterns:
        match = re.search(pattern, name)
        if match:
            number = int(match.group(1))
            if number < 1:
                raise ValueError(f"Numéro de tour invalide dans {path.name!r}.")
            return f"Tour {number}", number
    raise ValueError(
        f"Nom de fichier non reconnu : {path.name}. "
        "Utilisez par exemple 'Tour 3.xlsx', 'T3.csv' ou 'Tour Def.xlsx'."
    )


def as_int(value: Any, *, context: str) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ValueError(f"{context} : rang booléen invalide {value!r}.")
    if isinstance(value, (int, float)):
        if value != value:  # NaN
            return None
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("\u202f", "").replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return int(float(s))
    except ValueError as exc:
        raise ValueError(f"{context} : rang non numérique {value!r}.") from exc


def header_kind(value: Any) -> str | None:
    h = norm(value)
    if not h:
        return None
    if h in {"specialite", "speciality", "specialty", "discipline", "des", "des specialite"}:
        return "specialty"
    if "specialite" in h or "specialty" in h:
        return "specialty"
    if "ville" in h or "subdivision" in h or h in {"chu", "centre", "centre hospitalier"}:
        return "city"
    if h in {"rang max", "rang maximum", "rang maximal", "rang limite max", "rang limite maximal"}:
        return "max"
    if "rang" in h and any(word in h.split() for word in ("max", "maximum", "maximal")):
        return "max"
    if h in {"rang min", "rang minimum", "rang minimal"}:
        return "min"
    if "rang" in h and any(word in h.split() for word in ("min", "minimum", "minimal")):
        return "min"
    if h in {"rangs limites", "rang limite", "intervalle", "plage de rangs", "plage rangs"}:
        return "ranks"
    return None


def locate_header(rows: list[list[Any]], *, source: str, require_specialty: bool) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        found: dict[str, int] = {}
        for col_index, value in enumerate(row):
            kind = header_kind(value)
            if kind and kind not in found:
                found[kind] = col_index
        required = {"city", "max"} | ({"specialty"} if require_specialty else set())
        if required.issubset(found):
            return row_index, found
    requirement = "spécialité + ville/subdivision + rang max" if require_specialty else "ville/subdivision + rang max"
    preview = rows[: min(5, len(rows))]
    raise ValueError(f"{source} : en-tête introuvable ({requirement}). Premières lignes : {preview!r}")


def row_value(row: list[Any], index: int | None) -> Any:
    return row[index] if index is not None and index < len(row) else None


def parse_rows(
    rows: list[list[Any]], *, source: str, specialty_from_source: str | None = None
) -> dict[str, list[dict[str, Any]]]:
    require_specialty = specialty_from_source is None
    header_index, columns = locate_header(rows, source=source, require_specialty=require_specialty)
    out: "OrderedDict[str, list[dict[str, Any]]]" = OrderedDict()
    seen: dict[tuple[str, str], tuple[int | None, int | None, str]] = {}

    for physical_row, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(v not in (None, "") for v in row):
            continue
        raw_specialty = specialty_from_source if specialty_from_source is not None else row_value(row, columns.get("specialty"))
        raw_city = row_value(row, columns.get("city"))
        if raw_specialty in (None, "") and raw_city in (None, ""):
            continue
        if raw_specialty in (None, ""):
            raise ValueError(f"{source}, ligne {physical_row} : spécialité manquante.")
        if raw_city in (None, ""):
            # Ignore truly empty trailing/formatting lines, but not a row that contains a rank.
            if row_value(row, columns.get("max")) in (None, ""):
                continue
            raise ValueError(f"{source}, ligne {physical_row} : ville/subdivision manquante.")

        specialty = canonical_specialty(raw_specialty)
        city = canonical_city(raw_city)
        key = (norm(specialty), norm(city))

        maximum = as_int(row_value(row, columns.get("max")), context=f"{source}, ligne {physical_row}")
        minimum = as_int(row_value(row, columns.get("min")), context=f"{source}, ligne {physical_row}") if "min" in columns else None
        ranks = row_value(row, columns.get("ranks")) if "ranks" in columns else None
        ranks_text = "" if ranks is None else str(ranks).strip()

        # Une ligne strictement répétée n'altère pas les données finales : on l'ignore.
        # En revanche, deux lignes pour la même spécialité/ville avec des valeurs
        # différentes restent une erreur explicite.
        current = (maximum, minimum, ranks_text)
        if key in seen:
            if seen[key] == current:
                continue
            raise ValueError(f"{source}, ligne {physical_row} : doublon contradictoire pour {specialty} / {city}.")
        seen[key] = current
        if maximum is not None and maximum < 0:
            raise ValueError(f"{source}, ligne {physical_row} : rang max négatif ({maximum}).")
        if minimum is not None and minimum < 0:
            raise ValueError(f"{source}, ligne {physical_row} : rang min négatif ({minimum}).")
        if minimum is not None and maximum is not None and minimum > maximum:
            raise ValueError(f"{source}, ligne {physical_row} : rang min {minimum} > rang max {maximum}.")

        out.setdefault(specialty, []).append({
            "city": city,
            "max": maximum,
            "ranks": ranks_text,
        })
    return out


def sheet_rows(ws: Any) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def parse_xlsx(path: Path) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result: "OrderedDict[str, list[dict[str, Any]]]" = OrderedDict()
    try:
        if not wb.sheetnames:
            raise ValueError(f"{path.name} : classeur sans onglet.")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            specialty = canonical_specialty(sheet_name)
            parsed = parse_rows(sheet_rows(ws), source=f"{path.name} / {sheet_name}", specialty_from_source=specialty)
            # A worksheet represents one specialty. If the canonicalized title collides
            # with another worksheet, fail rather than overwrite data.
            for parsed_specialty, rows in parsed.items():
                if parsed_specialty in result:
                    raise ValueError(f"{path.name} : deux onglets correspondent à la spécialité {parsed_specialty!r}.")
                result[parsed_specialty] = rows
        if not result:
            raise ValueError(f"{path.name} : aucune donnée exploitable.")
        return result
    finally:
        wb.close()


def detect_csv_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        class Semicolon(csv.excel):
            delimiter = ";"
        return Semicolon()


def read_csv_rows(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError(f"{path.name} : encodage CSV non reconnu.")
    dialect = detect_csv_dialect(text[:8192])
    return [list(row) for row in csv.reader(text.splitlines(), dialect)]


def parse_csv(path: Path) -> dict[str, list[dict[str, Any]]]:
    rows = read_csv_rows(path)
    if not rows:
        raise ValueError(f"{path.name} : CSV vide.")
    parsed = parse_rows(rows, source=path.name, specialty_from_source=None)
    if not parsed:
        raise ValueError(f"{path.name} : aucune donnée exploitable.")
    return parsed


def parse_source(path: Path) -> dict[str, list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return parse_xlsx(path)
    if suffix == ".csv":
        return parse_csv(path)
    raise ValueError(f"Format non pris en charge : {path.name}")


def source_files(year_dir: Path) -> list[Path]:
    return sorted(
        p for p in year_dir.iterdir()
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS and not p.name.startswith("~$")
    )


def build_year(year_dir: Path) -> int:
    year = int(year_dir.name)
    files = source_files(year_dir)
    if not files:
        return 0

    labelled: list[tuple[int, str, Path]] = []
    seen_tours: dict[str, str] = {}
    for path in files:
        label, order = tour_from_filename(path)
        if label in seen_tours:
            raise ValueError(
                f"Deux fichiers correspondent à {label} dans {year_dir}: "
                f"{seen_tours[label]} et {path.name}. Gardez un seul fichier par tour."
            )
        seen_tours[label] = path.name
        labelled.append((order, label, path))
    labelled.sort(key=lambda item: item[0])

    # specialty key -> canonical display name, preserving first occurrence order
    specialty_display: "OrderedDict[str, str]" = OrderedDict()
    # display name -> tour -> rows
    merged: dict[str, dict[str, list[dict[str, Any]]]] = {}
    city_display: "OrderedDict[str, str]" = OrderedDict()
    tours: list[str] = []

    for _, label, path in labelled:
        print(f"[{year}] {label} <- {path.name}")
        parsed = parse_source(path)
        tours.append(label)
        seen_specialties_this_tour: set[str] = set()

        for raw_specialty, rows in parsed.items():
            skey = norm(raw_specialty)
            display = specialty_display.setdefault(skey, canonical_specialty(raw_specialty))
            if skey in seen_specialties_this_tour:
                raise ValueError(f"{path.name} : spécialité dupliquée après normalisation : {display}.")
            seen_specialties_this_tour.add(skey)
            merged.setdefault(display, {})[label] = rows
            for row in rows:
                ckey = norm(row["city"])
                canonical = city_display.setdefault(ckey, canonical_city(row["city"]))
                row["city"] = canonical

    # Make every specialty explicitly contain every tour. Missing specialty in a
    # particular round becomes an empty array, not a reused/old value.
    specialty_order = list(specialty_display.values())
    for specialty in specialty_order:
        merged.setdefault(specialty, {})
        for tour in tours:
            merged[specialty].setdefault(tour, [])

    payload = {
        "schema_version": 2,
        "year": year,
        "tours": tours,
        "specialty_order": specialty_order,
        "city_order": sorted(city_display.values(), key=lambda s: norm(s)),
        "specialties": merged,
    }
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    target = DATA_DIR / f"{year}.json"
    target.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Écrit : {target.relative_to(ROOT)} ({len(tours)} étapes, {len(specialty_order)} spécialités)")
    return len(tours)


def update_manifest() -> None:
    years = sorted(int(p.stem) for p in DATA_DIR.glob("*.json") if p.stem.isdigit())
    (DATA_DIR / "manifest.json").write_text(
        json.dumps({"schema_version": 2, "years": years}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print("Années publiées :", ", ".join(map(str, years)))


def iter_year_dirs() -> Iterable[Path]:
    if not SOURCES.exists():
        return []
    return sorted(
        (p for p in SOURCES.iterdir() if p.is_dir() and re.fullmatch(r"20\d{2}", p.name)),
        key=lambda p: int(p.name),
    )


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    for year_dir in iter_year_dirs():
        build_year(year_dir)
    update_manifest()


if __name__ == "__main__":
    main()
